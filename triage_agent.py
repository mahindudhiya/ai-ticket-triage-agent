"""
AI Ticket Triage Agent
======================
Reads a Zendesk-style CSV, classifies each ticket using the Claude API,
and outputs a formatted Excel dashboard with summary metrics.

Usage:
    1. Install dependencies:  pip install anthropic pandas openpyxl
    2. Set your API key:       export ANTHROPIC_API_KEY="your-key-here"
    3. Run:                    python triage_agent.py
"""

import os
import csv
import json
import time
import anthropic
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Config ──────────────────────────────────────────────────────────────────
INPUT_FILE  = "mock_tickets.csv"
OUTPUT_CSV  = "triage_output.csv"
OUTPUT_XLSX = "triage_dashboard.xlsx"
MODEL       = "claude-haiku-4-5-20251001"   # cheapest, fast — well under $1 for 60 tickets
# ─────────────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are a support operations analyst for Wealthsimple, a Canadian fintech company.
Your job is to classify inbound support tickets so the team can route, prioritize, and respond efficiently.

For each ticket, return ONLY a valid JSON object with these exact keys:
{
  "issue_type": one of ["Account Access", "Billing & Payments", "Technical Bug", "KYC & Compliance", "Portfolio & Trading", "Product Education", "Feature Request", "Security"],
  "urgency": one of ["Low", "Medium", "High", "Critical"],
  "sentiment": one of ["Positive", "Neutral", "Frustrated", "Angry"],
  "can_be_automated": true or false,
  "recommended_macro": a short macro name like "Reset Password Flow" or "Duplicate Charge Refund Process",
  "one_line_summary": a single sentence describing the ticket for an agent dashboard,
  "estimated_handle_time_mins": an integer estimate of how long this ticket would take a human agent (2–20)
}

Rules:
- Critical = financial loss, security breach, account locked with time-sensitive need
- can_be_automated = true if a standard macro or bot flow could fully resolve this without human judgement
- Be consistent. Use the exact strings listed above for issue_type, urgency, and sentiment.
- Return ONLY the JSON. No explanation, no markdown, no preamble."""


def classify_ticket(client: anthropic.Anthropic, ticket: dict) -> dict:
    """Send one ticket to Claude and return parsed classification."""
    user_message = f"""Ticket #{ticket['ticket_id']}
Subject: {ticket['subject']}
Description: {ticket['description']}
Channel: {ticket['channel']}
Priority (user-reported): {ticket['priority']}"""

    try:
        response = client.messages.create(
            model=MODEL,
            max_tokens=300,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_message}]
        )
        raw = response.content[0].text.strip()
        return json.loads(raw)
    except json.JSONDecodeError:
        # Fallback if model returns non-JSON
        return {
            "issue_type": "Technical Bug",
            "urgency": "Medium",
            "sentiment": "Neutral",
            "can_be_automated": False,
            "recommended_macro": "Manual Review Required",
            "one_line_summary": "Classification failed — manual review needed.",
            "estimated_handle_time_mins": 10
        }


def run_triage():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise EnvironmentError("ANTHROPIC_API_KEY not set. Run: export ANTHROPIC_API_KEY='your-key'")

    client = anthropic.Anthropic(api_key=api_key)

    # Load tickets
    with open(INPUT_FILE, newline="", encoding="utf-8") as f:
        tickets = list(csv.DictReader(f))

    print(f"Loaded {len(tickets)} tickets. Starting classification...\n")

    results = []
    for i, ticket in enumerate(tickets, 1):
        print(f"  [{i:02d}/{len(tickets)}] Ticket #{ticket['ticket_id']} — {ticket['subject'][:55]}...")
        classification = classify_ticket(client, ticket)
        combined = {**ticket, **classification}
        results.append(combined)
        time.sleep(0.3)  # polite rate limit

    print(f"\n✅ All {len(tickets)} tickets classified.")

    # Save raw output CSV
    df = pd.DataFrame(results)
    df.to_csv(OUTPUT_CSV, index=False)
    print(f"📄 Raw output saved: {OUTPUT_CSV}")

    # Build Excel dashboard
    build_dashboard(df)
    print(f"📊 Dashboard saved:  {OUTPUT_XLSX}")
    print("\n─────────────────────────────────────")
    print_summary(df)


def print_summary(df: pd.DataFrame):
    total = len(df)
    automatable = df["can_be_automated"].sum()
    containment_rate = automatable / total * 100
    time_saved = df[df["can_be_automated"] == True]["estimated_handle_time_mins"].sum()
    critical_count = (df["urgency"] == "Critical").sum()
    print(f"SUMMARY METRICS")
    print(f"  Total tickets analyzed : {total}")
    print(f"  Automatable tickets    : {automatable} ({containment_rate:.0f}% containment rate)")
    print(f"  Estimated time saved   : {time_saved} mins ({time_saved/60:.1f} hrs) per batch")
    print(f"  Critical tickets       : {critical_count}")
    print("─────────────────────────────────────")


# ── Excel Dashboard Builder ───────────────────────────────────────────────────

COLORS = {
    "header_dark":  "1A1A2E",
    "header_text":  "FFFFFF",
    "accent_blue":  "0077B6",
    "accent_green": "2DC653",
    "accent_red":   "E63946",
    "accent_yellow":"FFC300",
    "row_alt":      "F0F4F8",
    "white":        "FFFFFF",
    "light_gray":   "E8ECEF",
}

def hdr(ws, row, col, value, bg=None, bold=True, color="000000", size=11, wrap=False, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.font = Font(bold=bold, color=color, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return cell

def add_border(ws, min_row, max_row, min_col, max_col, color="CCCCCC"):
    thin = Side(style="thin", color=color)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def build_dashboard(df: pd.DataFrame):
    wb = Workbook()

    # ── Sheet 1: Executive Summary ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Executive Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 3

    # Title banner
    ws1.merge_cells("B2:K3")
    title_cell = ws1["B2"]
    title_cell.value = "AI TICKET TRIAGE AGENT — EXECUTIVE SUMMARY"
    title_cell.fill = PatternFill("solid", start_color=COLORS["header_dark"])
    title_cell.font = Font(bold=True, color=COLORS["header_text"], size=16, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws1.merge_cells("B4:K4")
    sub = ws1["B4"]
    sub.value = f"Wealthsimple Support Operations  |  Batch processed: {datetime.now().strftime('%B %d, %Y')}  |  Model: Claude Haiku"
    sub.font = Font(color="666666", size=10, italic=True, name="Calibri")
    sub.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 30
    ws1.row_dimensions[3].height = 30

    # KPI cards
    total = len(df)
    automatable = int(df["can_be_automated"].sum())
    containment_pct = f"{automatable/total*100:.0f}%"
    time_saved_hrs = f"{df[df['can_be_automated']==True]['estimated_handle_time_mins'].sum()/60:.1f} hrs"
    critical = int((df["urgency"]=="Critical").sum())
    avg_handle = f"{df['estimated_handle_time_mins'].mean():.1f} min"

    kpis = [
        ("Total Tickets", str(total),       COLORS["accent_blue"],  "Batch analyzed"),
        ("Containment Rate", containment_pct, COLORS["accent_green"], "Auto-resolvable"),
        ("Time Saved", time_saved_hrs,        "7B2D8B",               "Per batch"),
        ("Critical Tickets", str(critical),   COLORS["accent_red"],   "Need immediate action"),
        ("Avg Handle Time", avg_handle,       "E07B39",               "Per ticket"),
    ]
    kpi_cols = [2, 4, 6, 8, 10]
    ws1.row_dimensions[6].height = 20
    ws1.row_dimensions[7].height = 50
    ws1.row_dimensions[8].height = 25
    ws1.row_dimensions[9].height = 20

    for col, (label, value, color, sub_label) in zip(kpi_cols, kpis):
        ws1.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
        ws1.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)
        ws1.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+1)
        ws1.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col+1)
        c_label = ws1.cell(row=6, column=col, value=label.upper())
        c_label.fill = PatternFill("solid", start_color=color)
        c_label.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        c_label.alignment = Alignment(horizontal="center", vertical="center")
        c_val = ws1.cell(row=7, column=col, value=value)
        c_val.fill = PatternFill("solid", start_color=color)
        c_val.font = Font(bold=True, color="FFFFFF", size=26, name="Calibri")
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_sub = ws1.cell(row=8, column=col, value=sub_label)
        c_sub.fill = PatternFill("solid", start_color=color)
        c_sub.font = Font(color="DDDDDD", size=9, name="Calibri")
        c_sub.alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=9, column=col).fill = PatternFill("solid", start_color=color)
        for r in [6,7,8,9]:
            ws1.cell(row=r, column=col+1).fill = PatternFill("solid", start_color=color)

    ws1.row_dimensions[10].height = 10

    # Breakdown tables
    def breakdown_table(ws, start_row, start_col, title, series, bg_color):
        ws.row_dimensions[start_row].height = 22
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+2)
        t = ws.cell(row=start_row, column=start_col, value=title)
        t.fill = PatternFill("solid", start_color=bg_color)
        t.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        t.alignment = Alignment(horizontal="center", vertical="center")
        hdr(ws, start_row+1, start_col,   "Category",  bg=COLORS["light_gray"], align="left")
        hdr(ws, start_row+1, start_col+1, "Count",     bg=COLORS["light_gray"])
        hdr(ws, start_row+1, start_col+2, "% of Total",bg=COLORS["light_gray"])
        add_border(ws, start_row+1, start_row+1, start_col, start_col+2)
        for i, (cat, cnt) in enumerate(series.items()):
            r = start_row+2+i
            fill = COLORS["row_alt"] if i % 2 == 0 else COLORS["white"]
            ws.cell(row=r, column=start_col, value=cat).fill = PatternFill("solid", start_color=fill)
            ws.cell(row=r, column=start_col, value=cat).font = Font(name="Calibri", size=10)
            ws.cell(row=r, column=start_col, value=cat).alignment = Alignment(horizontal="left")
            ws.cell(row=r, column=start_col+1, value=int(cnt)).fill = PatternFill("solid", start_color=fill)
            ws.cell(row=r, column=start_col+1, value=int(cnt)).alignment = Alignment(horizontal="center")
            pct_val = f"{cnt/total*100:.0f}%"
            ws.cell(row=r, column=start_col+2, value=pct_val).fill = PatternFill("solid", start_color=fill)
            ws.cell(row=r, column=start_col+2, value=pct_val).alignment = Alignment(horizontal="center")
        add_border(ws, start_row+2, start_row+1+len(series), start_col, start_col+2)

    issue_counts    = df["issue_type"].value_counts().to_dict()
    urgency_counts  = df["urgency"].value_counts().to_dict()
    sentiment_counts= df["sentiment"].value_counts().to_dict()

    breakdown_table(ws1, 11, 2, "TICKETS BY ISSUE TYPE",  issue_counts,    COLORS["accent_blue"])
    breakdown_table(ws1, 11, 6, "TICKETS BY URGENCY",     urgency_counts,  COLORS["accent_red"])
    breakdown_table(ws1, 11, 9, "TICKETS BY SENTIMENT",   sentiment_counts,"7B2D8B")

    for col_letter, width in [("B",25),("C",8),("D",12),("E",3),("F",20),("G",8),("H",12),("I",3),("J",20),("K",8),("L",12)]:
        ws1.column_dimensions[col_letter].width = width

    # ── Sheet 2: All Tickets ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("🎫 All Tickets")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"

    cols = [
        ("Ticket ID", 10), ("Date", 16), ("Subject", 40), ("Channel", 10),
        ("Issue Type", 20), ("Urgency", 12), ("Sentiment", 13),
        ("Automatable?", 14), ("Macro / Next Step", 30),
        ("Summary", 45), ("Handle Time (min)", 18)
    ]
    for c, (label, width) in enumerate(cols, 1):
        cell = ws2.cell(row=1, column=c, value=label)
        cell.fill = PatternFill("solid", start_color=COLORS["header_dark"])
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(c)].width = width
    ws2.row_dimensions[1].height = 32

    urgency_colors = {"Critical": "FFCCCC","High": "FFE5CC","Medium": "FFFACC","Low": "E5FFCC"}
    sentiment_colors = {"Angry":"FFCCCC","Frustrated":"FFE5CC","Neutral":"F0F4F8","Positive":"E5FFCC"}

    for i, row in df.iterrows():
        r = i + 2
        fill = COLORS["row_alt"] if i % 2 == 0 else COLORS["white"]
        auto_str = "✅ Yes" if row.get("can_be_automated") else "❌ No"
        values = [
            row["ticket_id"], row["created_at"], row["subject"], row["channel"],
            row.get("issue_type",""), row.get("urgency",""), row.get("sentiment",""),
            auto_str, row.get("recommended_macro",""),
            row.get("one_line_summary",""), row.get("estimated_handle_time_mins","")
        ]
        for c, val in enumerate(values, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = Alignment(vertical="center", wrap_text=(c in [3,9,10]))
            if c == 6:
                urg = str(row.get("urgency",""))
                bg = urgency_colors.get(urg, fill)
                cell.fill = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 7:
                snt = str(row.get("sentiment",""))
                bg = sentiment_colors.get(snt, fill)
                cell.fill = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = PatternFill("solid", start_color=fill)
        ws2.row_dimensions[r].height = 28

    add_border(ws2, 1, len(df)+1, 1, len(cols))

    # ── Sheet 3: Critical Tickets ────────────────────────────────────────────
    ws3 = wb.create_sheet("🚨 Critical & High")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"

    critical_df = df[df["urgency"].isin(["Critical","High"])].copy()

    ws3.merge_cells("A1:K1")
    banner = ws3["A1"]
    banner.value = f"⚠️  CRITICAL & HIGH PRIORITY TICKETS — {len(critical_df)} tickets require urgent attention"
    banner.fill = PatternFill("solid", start_color=COLORS["accent_red"])
    banner.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    short_cols = [
        ("Ticket ID",10),("Subject",40),("Description",55),
        ("Urgency",12),("Issue Type",20),("Sentiment",13),
        ("Macro",30),("Handle Time",14)
    ]
    for c, (label, width) in enumerate(short_cols, 1):
        cell = ws3.cell(row=2, column=c, value=label)
        cell.fill = PatternFill("solid", start_color="B00020")
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws3.column_dimensions[get_column_letter(c)].width = width
    ws3.row_dimensions[2].height = 28

    for i, (_, row) in enumerate(critical_df.iterrows()):
        r = i + 3
        bg = "FFE5E5" if i % 2 == 0 else "FFF5F5"
        vals = [
            row["ticket_id"], row["subject"], row["description"],
            row.get("urgency",""), row.get("issue_type",""), row.get("sentiment",""),
            row.get("recommended_macro",""), row.get("estimated_handle_time_mins","")
        ]
        for c, val in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = Alignment(vertical="center", wrap_text=(c in [2,3,7]))
        ws3.row_dimensions[r].height = 40

    add_border(ws3, 2, 2+len(critical_df), 1, len(short_cols))

    # ── Sheet 4: Automation Candidates ──────────────────────────────────────
    ws4 = wb.create_sheet("🤖 Automation Candidates")
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = "A2"

    auto_df = df[df["can_be_automated"] == True].copy()

    ws4.merge_cells("A1:I1")
    banner4 = ws4["A1"]
    banner4.value = f"🤖  AUTOMATION CANDIDATES — {len(auto_df)} tickets ({len(auto_df)/total*100:.0f}% containment rate) — Est. {auto_df['estimated_handle_time_mins'].sum()} mins saved per batch"
    banner4.fill = PatternFill("solid", start_color=COLORS["accent_green"])
    banner4.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    banner4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30

    auto_cols = [
        ("Ticket ID",10),("Subject",40),("Issue Type",20),
        ("Urgency",12),("Macro / Automation Flow",35),("Handle Time (min)",16)
    ]
    for c, (label, width) in enumerate(auto_cols, 1):
        cell = ws4.cell(row=2, column=c, value=label)
        cell.fill = PatternFill("solid", start_color="1A6B3C")
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws4.column_dimensions[get_column_letter(c)].width = width
    ws4.row_dimensions[2].height = 28

    for i, (_, row) in enumerate(auto_df.iterrows()):
        r = i + 3
        bg = "E8F8EE" if i % 2 == 0 else "F5FFF8"
        vals = [
            row["ticket_id"], row["subject"], row.get("issue_type",""),
            row.get("urgency",""), row.get("recommended_macro",""),
            row.get("estimated_handle_time_mins","")
        ]
        for c, val in enumerate(vals, 1):
            cell = ws4.cell(row=r, column=c, value=val)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = Alignment(vertical="center", wrap_text=(c in [2,5]))
        ws4.row_dimensions[r].height = 25

    add_border(ws4, 2, 2+len(auto_df), 1, len(auto_cols))

    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    run_triage()
